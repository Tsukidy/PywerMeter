# This script handles the excel file operations for the pywerHelper package.
# It includes functions to read from and write to Excel files using the pandas library.
# Author: Dylan Pope!
# Date: 2025-11-07
# Version: 1.0.0

import os
import logging
import sys

# Set up module logger
logger = logging.getLogger(__name__)
logger.setLevel(logging.DEBUG)

# Try to import pandas and openpyxl with helpful error messages
try:
    import pandas as pd
    logger.info("pandas imported successfully")
except ImportError as e:
    logger.error(f"pandas is not installed: {e}")
    print("ERROR: pandas is not installed. Install with: pip install pandas")
    raise

try:
    import openpyxl
    logger.info("openpyxl imported successfully")
except ImportError as e:
    logger.error(f"openpyxl is not installed: {e}")
    print("ERROR: openpyxl is not installed. Install with: pip install openpyxl")
    raise


def testFunction():
    """Test function to verify module import."""
    return "This is a test function from excelHelper.py"


def create_workbook(output_filename="power_data.xlsx"):
    """
    Create a new Excel workbook with an empty DataFrame.
    
    Args:
        output_filename (str): The filename for the new Excel workbook
        
    Returns:
        str: Path to the created workbook, or None if failed
    """
    try:
        logger.info(f"Creating new workbook: {output_filename}")
        # Create an empty DataFrame
        df = pd.DataFrame()
        
        # Write to Excel
        with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Summary', index=False)
        
        logger.info(f"Successfully created new workbook: {output_filename}")
        print(f"Created new workbook: {output_filename}")
        return output_filename
    except PermissionError:
        logger.error(f"Permission denied creating workbook: {output_filename}", exc_info=True)
        print(f"ERROR: Permission denied. File may be open in another program: {output_filename}")
        return None
    except OSError as e:
        logger.error(f"OS error creating workbook: {e}", exc_info=True)
        print(f"ERROR: Failed to create workbook: {e}")
        return None
    except Exception as e:
        logger.error(f"Unexpected error creating workbook: {e}", exc_info=True)
        print(f"ERROR: Unexpected error creating workbook: {e}")
        return None


def import_data_to_workbook(data_file, workbook_filename="power_data.xlsx", sheet_name=None):
    """
    Import data from a text file into an Excel workbook as a new sheet.
    Each line from the data file becomes a row in the worksheet.
    
    Args:
        data_file (str): Path to the data file to import
        workbook_filename (str): Path to the Excel workbook
        sheet_name (str): Name for the sheet. If None, uses the data file name without extension
        
    Returns:
        bool: True if successful, False otherwise
    """
    try:
        logger.info(f"Importing data from {data_file} to {workbook_filename}")
        
        # Check if data file exists
        if not os.path.exists(data_file):
            logger.warning(f"Data file not found: {data_file}")
            print(f"ERROR: Data file not found: {data_file}")
            return False
        
        # Read data from text file
        try:
            with open(data_file, 'r', encoding='utf-8') as f:
                lines = f.readlines()
        except PermissionError:
            logger.error(f"Permission denied reading file: {data_file}", exc_info=True)
            print(f"ERROR: Permission denied reading file: {data_file}")
            return False
        except UnicodeDecodeError as e:
            logger.error(f"Unicode decode error reading {data_file}: {e}", exc_info=True)
            print(f"ERROR: File encoding issue: {e}")
            return False
        
        # Strip newlines and create DataFrame
        data = [line.strip() for line in lines if line.strip()]
        
        if not data:
            logger.warning(f"No data found in file: {data_file}")
            print(f"WARNING: No data found in file: {data_file}")
            return False
        
        df = pd.DataFrame({'Power Data': data})
        
        # Determine sheet name
        if sheet_name is None:
            sheet_name = os.path.splitext(os.path.basename(data_file))[0]
        
        # Sanitize sheet name (Excel has 31 char limit and special char restrictions)
        sheet_name = sheet_name[:31].replace('[', '').replace(']', '').replace('*', '').replace('/', '').replace('\\', '').replace('?', '').replace(':', '')
        
        # Check if workbook exists
        if os.path.exists(workbook_filename):
            # Append to existing workbook
            try:
                with pd.ExcelWriter(workbook_filename, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
            except PermissionError:
                logger.error(f"Permission denied writing to workbook: {workbook_filename}", exc_info=True)
                print(f"ERROR: Permission denied. Workbook may be open: {workbook_filename}")
                return False
        else:
            # Create new workbook
            try:
                with pd.ExcelWriter(workbook_filename, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
            except PermissionError:
                logger.error(f"Permission denied creating workbook: {workbook_filename}", exc_info=True)
                print(f"ERROR: Permission denied creating workbook: {workbook_filename}")
                return False
        
        logger.info(f"Successfully imported {len(data)} rows from {data_file} to sheet '{sheet_name}'")
        print(f"Imported {len(data)} rows from {data_file} to sheet '{sheet_name}'")
        return True
        
    except ValueError as e:
        logger.error(f"Value error importing data from {data_file}: {e}", exc_info=True)
        print(f"ERROR: Invalid data format: {e}")
        return False
    except Exception as e:
        logger.error(f"Unexpected error importing data from {data_file}: {e}", exc_info=True)
        print(f"ERROR: Unexpected error importing data: {e}")
        return False


def import_multiple_files(data_files, workbook_filename="power_data.xlsx"):
    """
    Import multiple data files into a single Excel workbook, each as a separate sheet.
    
    Args:
        data_files (list): List of data file paths to import
        workbook_filename (str): Path to the Excel workbook
        
    Returns:
        int: Number of files successfully imported
    """
    if not data_files:
        logger.warning("No data files provided to import")
        print("WARNING: No data files provided to import")
        return 0
    
    logger.info(f"Importing {len(data_files)} files to {workbook_filename}")
    success_count = 0
    
    for data_file in data_files:
        if import_data_to_workbook(data_file, workbook_filename):
            success_count += 1
    
    logger.info(f"Successfully imported {success_count} out of {len(data_files)} files to {workbook_filename}")
    print(f"\nImported {success_count} out of {len(data_files)} files to {workbook_filename}")
    return success_count


def write_test_row_to_excel(test_header, samples, workbook_filename="power_measurements.xlsx", sheet_name="Power Data"):
    """
    Write a test's data as a new column in an Excel file.
    
    Args:
        test_header (str): Name of the test (goes in column header)
        samples (list): List of sample values for this test
        workbook_filename (str): Path to the Excel workbook
        sheet_name (str): Name for the sheet
        
    Returns:
        bool: True if successful, False otherwise
    """
    try:
        logger.info(f"Writing test '{test_header}' with {len(samples)} samples to {workbook_filename}")
        
        # Convert samples to numeric values (handles text strings from serial data)
        try:
            numeric_samples = pd.to_numeric(samples, errors='coerce')
            # Count how many values failed to convert (pd.isna works with numpy arrays)
            nan_count = pd.isna(numeric_samples).sum()
            if nan_count > 0:
                logger.warning(f"{nan_count} samples could not be converted to numeric values")
        except Exception as e:
            logger.error(f"Error converting samples to numeric: {e}", exc_info=True)
            print(f"ERROR: Failed to process sample data: {e}")
            return False
        
        # Check if workbook exists
        if os.path.exists(workbook_filename):
            # Read existing data
            try:
                existing_df = pd.read_excel(workbook_filename, sheet_name=sheet_name)
                logger.debug(f"Loaded existing workbook with {len(existing_df)} rows")
                
                # Ensure DataFrame has enough rows for new samples
                if len(samples) > len(existing_df):
                    # Add empty rows
                    rows_to_add = len(samples) - len(existing_df)
                    empty_rows = pd.DataFrame([['' for _ in existing_df.columns] for _ in range(rows_to_add)], 
                                             columns=existing_df.columns)
                    existing_df = pd.concat([existing_df, empty_rows], ignore_index=True)
                    logger.debug(f"Added {rows_to_add} empty rows to accommodate new data")
                
                # Add new column with test data (as numeric values)
                existing_df[test_header] = numeric_samples
                updated_df = existing_df
                
            except FileNotFoundError:
                # Sheet doesn't exist, create new DataFrame
                logger.info(f"Sheet '{sheet_name}' not found, creating new DataFrame")
                updated_df = pd.DataFrame({test_header: numeric_samples})
            except PermissionError:
                logger.error(f"Permission denied reading workbook: {workbook_filename}", exc_info=True)
                print(f"ERROR: Permission denied. File may be open: {workbook_filename}")
                return False
            except ValueError as e:
                # Sheet doesn't exist or other pandas error
                logger.info(f"Creating new DataFrame due to: {e}")
                updated_df = pd.DataFrame({test_header: numeric_samples})
        else:
            # Create new workbook with first column
            logger.info(f"Creating new workbook: {workbook_filename}")
            updated_df = pd.DataFrame({test_header: numeric_samples})
        
        # Write to Excel without index and without default header names
        try:
            with pd.ExcelWriter(workbook_filename, engine='openpyxl', mode='w') as writer:
                updated_df.to_excel(writer, sheet_name=sheet_name, index=False)
        except PermissionError:
            logger.error(f"Permission denied writing to workbook: {workbook_filename}", exc_info=True)
            print(f"ERROR: Permission denied. File may be open: {workbook_filename}")
            return False
        except OSError as e:
            logger.error(f"OS error writing to workbook: {e}", exc_info=True)
            print(f"ERROR: Failed to write to file: {e}")
            return False
        
        logger.info(f"Successfully wrote test '{test_header}' to {workbook_filename}")
        print(f"Wrote test '{test_header}' to Excel: {len(samples)} samples")
        return True
        
    except Exception as e:
        logger.error(f"Unexpected error writing test column to Excel: {e}", exc_info=True)
        print(f"ERROR: Unexpected error writing to Excel: {e}")
        return False

class PowerCalc:
    """
    A class to handle power calculations and Excel formula operations.
    """
    
    def __init__(self, workbook_filename="power_measurements.xlsx", sheet_name="Power Data"):
        """
        Initialize the PowerCalc class.
        
        Args:
            workbook_filename (str): Path to the Excel workbook
            sheet_name (str): Name of the sheet to work with
        """
        self.workbook_filename = workbook_filename
        self.sheet_name = sheet_name
        self.wb = None
        self.ws = None
        self.df = None
        self.last_data_row = None
    
    def _load_workbook(self):
        """Load the Excel workbook and worksheet with comprehensive error handling."""
        try:
            if not os.path.exists(self.workbook_filename):
                logger.warning(f"Workbook not found: {self.workbook_filename}")
                print(f"ERROR: Workbook not found: {self.workbook_filename}")
                return False
            
            # Load workbook with openpyxl first to check structure
            try:
                from openpyxl import load_workbook
                self.wb = load_workbook(self.workbook_filename)
            except PermissionError:
                logger.error(f"Permission denied loading workbook with openpyxl: {self.workbook_filename}", exc_info=True)
                print(f"ERROR: Permission denied. File may be open: {self.workbook_filename}")
                return False
            except Exception as e:
                logger.error(f"Error loading workbook with openpyxl: {e}", exc_info=True)
                print(f"ERROR: Failed to load workbook: {e}")
                return False
            
            try:
                self.ws = self.wb[self.sheet_name]
            except KeyError:
                logger.error(f"Sheet '{self.sheet_name}' not found in workbook", exc_info=True)
                print(f"ERROR: Sheet '{self.sheet_name}' not found")
                return False
            
            # Check if averages are at the top by looking at cell A1
            has_averages_at_top = (self.ws['A1'].value == 'Averages')
            
            # Read the Excel file with pandas
            # If averages are at top, row 3 is the header; otherwise row 1
            try:
                if has_averages_at_top:
                    self.df = pd.read_excel(self.workbook_filename, sheet_name=self.sheet_name, header=2)
                    logger.debug("Detected averages at top, using row 3 as header")
                else:
                    self.df = pd.read_excel(self.workbook_filename, sheet_name=self.sheet_name)
                    logger.debug("No averages at top, using row 1 as header")
            except FileNotFoundError:
                logger.error(f"Workbook file not found: {self.workbook_filename}", exc_info=True)
                print(f"ERROR: File not found: {self.workbook_filename}")
                return False
            except PermissionError:
                logger.error(f"Permission denied accessing workbook: {self.workbook_filename}", exc_info=True)
                print(f"ERROR: Permission denied. File may be open: {self.workbook_filename}")
                return False
            except ValueError as e:
                logger.error(f"Sheet '{self.sheet_name}' not found in workbook: {e}", exc_info=True)
                print(f"ERROR: Sheet '{self.sheet_name}' not found in workbook")
                return False
            
            if self.df.empty:
                logger.warning(f"No data found in workbook sheet '{self.sheet_name}'")
                print(f"WARNING: No data found in workbook sheet '{self.sheet_name}'")
                return False
            
            # Find the last row with data
            self.last_data_row = len(self.df) + 1  # +1 for header row
            logger.info(f"Loaded workbook with {len(self.df)} data rows")
            
            return True
            
        except Exception as e:
            logger.error(f"Unexpected error loading workbook: {e}", exc_info=True)
            print(f"ERROR: Unexpected error loading workbook: {e}")
            return False
    
    def _save_and_close(self):
        """Save and close the workbook with error handling."""
        if self.wb:
            try:
                self.wb.save(self.workbook_filename)
                logger.info(f"Workbook saved: {self.workbook_filename}")
            except PermissionError:
                logger.error(f"Permission denied saving workbook: {self.workbook_filename}", exc_info=True)
                print(f"ERROR: Permission denied saving. File may be open: {self.workbook_filename}")
                raise
            except OSError as e:
                logger.error(f"OS error saving workbook: {e}", exc_info=True)
                print(f"ERROR: Failed to save workbook: {e}")
                raise
            finally:
                try:
                    self.wb.close()
                    logger.debug("Workbook closed")
                except Exception as e:
                    logger.warning(f"Error closing workbook: {e}")
    
    def add_averages(self):
        """
        Add average calculations to the top of each column in the Excel file with a spanning header.
        Restructures the sheet so that:
        Row 1: "Averages" header (spanning all columns)
        Row 2: Average formulas
        Row 3: Column headers
        Row 4+: Data
        
        Returns:
            bool: True if successful, False otherwise
        """
        try:
            logger.info(f"Adding averages to {self.workbook_filename}")
            if not self._load_workbook():
                return False
            
            from openpyxl.utils import get_column_letter
            
            # Check if averages already exist
            if self.ws['A1'].value == 'Averages':
                print("Averages already exist in this file. Skipping.")
                logger.warning("Averages already exist, skipping add_averages operation")
                self.wb.close()
                return True
            
            # Insert 2 new rows at the top
            self.ws.insert_rows(1, 2)
            logger.debug("Inserted 2 rows at the top for averages")
            
            # Add "Averages" header in row 1, column A, and merge across all columns
            num_cols = len(self.df.columns)
            self.ws['A1'] = 'Averages'
            if num_cols > 1:
                end_col_letter = get_column_letter(num_cols)
                self.ws.merge_cells(f'A1:{end_col_letter}1')
            
            # Style the Averages header (bold, centered, and bordered)
            from openpyxl.styles import Font, Alignment, Border, Side
            
            # Create border style
            border_style = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )
            
            self.ws['A1'].font = Font(bold=True)
            self.ws['A1'].alignment = Alignment(horizontal='center')
            self.ws['A1'].border = border_style
            
            # Apply border to all merged cells in row 1
            if num_cols > 1:
                for col_idx in range(1, num_cols + 1):
                    col_letter = get_column_letter(col_idx)
                    self.ws[f'{col_letter}1'].border = border_style
            
            # Add average formulas in row 2
            # Data now starts at row 4 (was row 2), goes to last_data_row + 2
            for col_idx, col_name in enumerate(self.df.columns, start=1):
                try:
                    col_letter = get_column_letter(col_idx)
                    
                    # Add AVERAGE formula in row 2
                    # Data range is now from row 4 to last_data_row + 2
                    formula_cell = self.ws[f'{col_letter}2']
                    formula_cell.value = f'=AVERAGE({col_letter}4:{col_letter}{self.last_data_row + 2})'
                    logger.debug(f"Added average formula for column {col_letter}")
                except Exception as e:
                    logger.error(f"Error adding average for column {col_idx}: {e}", exc_info=True)
                    print(f"WARNING: Failed to add average for column {col_name}")
            
            # Save the workbook
            try:
                self._save_and_close()
            except Exception as e:
                logger.error(f"Error saving workbook after adding averages: {e}", exc_info=True)
                print(f"ERROR: Failed to save changes: {e}")
                return False
            
            logger.info(f"Successfully added average calculations to {self.workbook_filename}")
            print(f"Added average calculations to Excel file")
            return True
            
        except Exception as e:
            logger.error(f"Unexpected error adding calculations: {e}", exc_info=True)
            print(f"ERROR: Unexpected error adding calculations: {e}")
            return False
    
    def totalAnnualPower(self):
        """
        Calculate and add Total Annual Power based on the formula:
        totalAnnualPower = 8760/1000*(off*0.15 + shortidle*0.45 + longidle*0.1 + sleep*0.3)
        
        Adds a "Total Annual Power" header to the right of the "Averages" span with the calculated value below it.
        
        Returns:
            bool: True if successful, False otherwise
        """
        try:
            logger.info(f"Adding Total Annual Power to {self.workbook_filename}")
            if not self._load_workbook():
                return False
            
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import Font, Alignment, Border, Side
            
            # Find columns by header name (case-insensitive, spaces removed)
            columns_needed = ['off', 'shortidle', 'longidle', 'sleep']
            column_positions = {}
            
            for col_idx, col_name in enumerate(self.df.columns, start=1):
                # Normalize column name: lowercase and remove spaces
                col_name_normalized = str(col_name).lower().replace(' ', '')
                if col_name_normalized in columns_needed:
                    column_positions[col_name_normalized] = col_idx
            
            # Check if all required columns exist
            missing_columns = [col for col in columns_needed if col not in column_positions]
            if missing_columns:
                logger.warning(f"Missing required columns for Total Annual Power: {missing_columns}")
                print(f"ERROR: Missing required columns: {missing_columns}")
                print(f"Required columns are: {', '.join(columns_needed)}")
                return False
            
            # Check if averages are at the top
            has_averages_at_top = (self.ws['A1'].value == 'Averages')
            
            if not has_averages_at_top:
                print("ERROR: Total Annual Power requires averages to be added first")
                logger.error("Attempted to add Total Annual Power without averages at top")
                return False
            
            # Check if Total Annual Power already exists by scanning row 1
            tap_col_idx = None
            tap_col_letter = None
            max_col = self.ws.max_column
            
            for col_idx in range(1, max_col + 1):
                col_letter = get_column_letter(col_idx)
                if self.ws[f'{col_letter}1'].value == 'Total Annual Power':
                    tap_col_idx = col_idx
                    tap_col_letter = col_letter
                    logger.info(f"Found existing Total Annual Power at column {col_letter}, will overwrite")
                    print("Total Annual Power already exists, updating values...")
                    break
            
            # If not found, place it in the column right after the last data column
            if tap_col_idx is None:
                num_cols = len(self.df.columns)
                tap_col_idx = num_cols + 1
                tap_col_letter = get_column_letter(tap_col_idx)
                logger.info(f"Adding new Total Annual Power column at {tap_col_letter}")
            
            # Create border style
            border_style = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )
            
            # Add "Total Annual Power" header in row 1
            self.ws[f'{tap_col_letter}1'] = 'Total Annual Power'
            self.ws[f'{tap_col_letter}1'].font = Font(bold=True)
            self.ws[f'{tap_col_letter}1'].alignment = Alignment(horizontal='center')
            self.ws[f'{tap_col_letter}1'].border = border_style
            
            # Set column width to fit the text "Total Annual Power" (approximately 20 characters)
            self.ws.column_dimensions[tap_col_letter].width = 22
            logger.debug(f"Set column {tap_col_letter} width to 22")
            
            # Get column letters for the formula
            # Need to account for the structure: row 1 = headers, row 2 = averages, row 3 = column names, row 4+ = data
            off_letter = get_column_letter(column_positions['off'])
            shortidle_letter = get_column_letter(column_positions['shortidle'])
            longidle_letter = get_column_letter(column_positions['longidle'])
            sleep_letter = get_column_letter(column_positions['sleep'])
            
            # Add the formula in row 2 (where the averages are)
            avg_row = 2
            formula = f'=8760/1000*({off_letter}{avg_row}*0.15+{sleep_letter}{avg_row}*0.45+{longidle_letter}{avg_row}*0.1+{shortidle_letter}{avg_row}*0.3)'
            self.ws[f'{tap_col_letter}{avg_row}'] = formula
            self.ws[f'{tap_col_letter}{avg_row}'].border = border_style
            logger.debug(f"Added Total Annual Power formula: {formula}")
            
            # Save the workbook
            try:
                self._save_and_close()
            except Exception as e:
                logger.error(f"Error saving workbook after adding Total Annual Power: {e}", exc_info=True)
                print(f"ERROR: Failed to save changes: {e}")
                return False
            
            logger.info(f"Successfully added Total Annual Power calculation to {self.workbook_filename}")
            print(f"Added Total Annual Power calculation to Excel file")
            return True
            
        except Exception as e:
            logger.error(f"Unexpected error adding Total Annual Power calculation: {e}", exc_info=True)
            print(f"ERROR: Unexpected error adding Total Annual Power calculation: {e}")
            return False


# Backward compatibility wrapper function
def powerCalc(workbook_filename="power_measurements.xlsx", sheet_name="Power Data"):
    """
    Legacy function wrapper for PowerCalc class.
    Add average calculations to the bottom of each column in the Excel file.
    
    Args:
        workbook_filename (str): Path to the Excel workbook
        sheet_name (str): Name of the sheet to add calculations to
        
    Returns:
        bool: True if successful, False otherwise
    """
    calc = PowerCalc(workbook_filename, sheet_name)
    return calc.add_averages()


# Module-level exports
__all__ = ['create_workbook', 'import_data_to_workbook', 'import_multiple_files', 'write_test_row_to_excel', 'powerCalc', 'PowerCalc']